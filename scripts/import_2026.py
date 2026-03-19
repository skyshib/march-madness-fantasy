#!/usr/bin/env python3
"""
Import 2026 March Madness Fantasy picks from Google Form xlsx export.

Reads the xlsx, normalizes player names via fuzzy matching against a canonical
player list built from all entries, resolves captains, and outputs picks.json.

Usage:
    python import_2026.py
"""

import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

try:
    import openpyxl
    from thefuzz import fuzz, process
except ImportError:
    print("Install dependencies: pip install openpyxl thefuzz")
    sys.exit(1)

# ── Paths ────────────────────────────────────────────────────────────────────
XLSX_PATH = Path.home() / "Downloads" / "March Madness Fantasy 2025 (Responses).xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "picks.json"

# ── Manual name corrections ──────────────────────────────────────────────────
# Map misspelling/variant (normalize_name_key form) -> canonical full name
NAME_OVERRIDES = {
    "cam boozer": "Cameron Boozer",
    "moncilovic": "Milan Momcilovic",
    "milan moncilovic": "Milan Momcilovic",
    "burries": "Brayden Burris",
    "brayden burries": "Brayden Burris",
    "dybanstsa": "AJ Dybantsa",
    "aj dybanstsa": "AJ Dybantsa",
    "aj dybansta": "AJ Dybantsa",
    "dybantsa": "AJ Dybantsa",
    "dybansta": "AJ Dybantsa",
    "skirtz": "Bennett Stirtz",
    "bennett skirtz": "Bennett Stirtz",
    "bennet stirtz": "Bennett Stirtz",
    "ejiofar": "Zuby Ejiofor",
    "zuby ejiofar": "Zuby Ejiofor",
    "thorton": "Bruce Thornton",
    "bruce thorton": "Bruce Thornton",
    "fleming": "Kingston Flemings",
    "kingston fleming": "Kingston Flemings",
    "tamari johnson": "Tavari Johnson",
    "labarlon philon": "Labaron Philon Jr.",
    "labarlon philon jr": "Labaron Philon Jr.",
    "labaron philon": "Labaron Philon Jr.",
    "labaron philon jr": "Labaron Philon Jr.",
    "terrace hill": "Terrence Hill Jr.",
    "terrace hill jr": "Terrence Hill Jr.",
    "terrence hill": "Terrence Hill Jr.",
    "terrence hill jr": "Terrence Hill Jr.",
    "rashaun agee": "Rashaun Agee",
    "rashawn agee": "Rashaun Agee",
    "donte horne": "Dontae Horne",
    "dontae home": "Dontae Horne",
    "otega owea": "Otega Oweh",
    "ja kobi gillespie": "Ja'Kobi Gillespie",
    "jakobi gillespie": "Ja'Kobi Gillespie",
    "dominique daniels": "Dominique Daniels Jr.",
    "dominique daniels jr": "Dominique Daniels Jr.",
    "dominique daniels": "Dominique Daniels Jr.",
    "mj collins": "MJ Collins Jr.",
    "mj collins jr": "MJ Collins Jr.",
    "victor valdes": "Victor Valdes",
    "darius acuff": "Darius Acuff Jr.",
    "darius acuff jr": "Darius Acuff Jr.",
    "augustas marciulionis": "Augustas Marciulionis",
    "taran armstrong": "Taran Armstrong",
    "marcus foster": "Marcus Foster",
    "jt toppin": "JT Toppin",
    "trey campbell": "Trey Campbell",
    "david punch": "David Punch",
    "chris youngblood": "Chris Youngblood",
    "bennett stirtz": "Bennett Stirtz",
    "lj cryer": "LJ Cryer",
    "caleb wilson": "Caleb Wilson",
    "jackson rasmussen": "Jackson Rasmussen",
    "travis harper": "Travis Harper II",
    "travis harper ii": "Travis Harper II",
    "dwt": "Damari Wheeler-Thomas",
    "damari wheeler thomas": "Damari Wheeler-Thomas",
    "damari wheeler-thomas": "Damari Wheeler-Thomas",
    "jamal fuller": "Jamal Fuller",
    "mason falslev": "Mason Falslev",
    "melvin council": "Melvin Council Jr.",
    "melvin council jr": "Melvin Council Jr.",
}

# ── Known team associations for players only mentioned without teams ─────────
# (e.g., Zack Duffy's entries have no teams)
PLAYER_TEAMS = {
    "JT Toppin": "Texas Tech",
    "Augustas Marciulionis": "Saint Mary's",
    "David Punch": "TCU",
    "Wade Taylor IV": "Texas A&M",
    "Chris Youngblood": "Texas",
    "Trey Campbell": "McNeese",
    "Taran Armstrong": "Troy",
    "Marcus Foster": "Queens",
    "LJ Cryer": "Houston",
    "Caleb Wilson": "BYU",
    "Jackson Rasmussen": "Idaho",
}

# ── Team name normalization ──────────────────────────────────────────────────
TEAM_OVERRIDES = {
    "isu": "Iowa State",
    "iowa st": "Iowa State",
    "iowa st.": "Iowa State",
    "msu": "Michigan State",
    "michigan st": "Michigan State",
    "michigan st.": "Michigan State",
    "usf": "South Florida",
    "tsu": "Tennessee State",
    "tennessee st": "Tennessee State",
    "tennessee st.": "Tennessee State",
    "ndsu": "North Dakota State",
    "north dakota st": "North Dakota State",
    "north dakota st.": "North Dakota State",
    "upenn": "Penn",
    "pennsylvania": "Penn",
    "ohio st": "Ohio State",
    "ohio st.": "Ohio State",
    "st johns": "St. John's",
    "st john's": "St. John's",
    "st. johns": "St. John's",
    "a&m": "Texas A&M",
    "texas a & m": "Texas A&M",
    "utah st": "Utah State",
    "utah st.": "Utah State",
    "utah state": "Utah State",
    "st marys": "Saint Mary's",
    "st. mary's": "Saint Mary's",
    "saint marys": "Saint Mary's",
    "saint mary's": "Saint Mary's",
    "vandy": "Vanderbilt",
    "miami": "Miami",
    "miami (fl)": "Miami",
    "miami (oh)": "Miami (OH)",
    "miami ohio": "Miami (OH)",
    "high point": "High Point",
    "cal baptist": "Cal Baptist",
    "california baptist": "Cal Baptist",
    "kennesaw": "Kennesaw State",
    "kennesaw st": "Kennesaw State",
    "kennesaw st.": "Kennesaw State",
    "sienna": "Siena",
    "siena": "Siena",
    "pv": "Prairie View A&M",
    "prairie view": "Prairie View A&M",
    "prairie view a & m": "Prairie View A&M",
    "prairie view a&m": "Prairie View A&M",
    "long island sharks": "LIU",
    "liu": "LIU",
    "howard bison": "Howard",
    "howard": "Howard",
    "mcneese state": "McNeese",
    "mcneese": "McNeese",
    "ucf": "UCF",
    "tennesee": "Tennessee",
    "tennessee": "Tennessee",
    "cal baptist": "Cal Baptist",
}

# ── Known school names for detecting "Name Team" without separator ───────────
KNOWN_SCHOOLS = {
    "duke", "houston", "illinois", "gonzaga", "arkansas", "kansas", "alabama",
    "vanderbilt", "tennessee", "byu", "ucla", "kentucky", "miami", "ohio state",
    "georgia", "iowa", "missouri", "texas", "south florida", "high point",
    "akron", "hofstra", "troy", "penn", "furman", "siena", "howard", "liu",
    "mcneese", "queens", "vcu", "villanova", "uconn", "purdue", "arizona",
    "michigan", "nebraska", "wisconsin", "florida", "iowa state",
    "michigan state", "louisville", "utah state", "santa clara", "texas a&m",
    "cal baptist", "kennesaw state", "north dakota state", "wright state",
    "tennessee state", "prairie view a&m", "hawaii", "idaho",
    "st. john's", "saint mary's",
}


def normalize_text(s):
    """Strip, collapse whitespace, normalize unicode."""
    if not s:
        return ""
    # Replace smart quotes/dashes
    s = s.replace("\u2018", "'").replace("\u2019", "'")
    s = s.replace("\u201c", '"').replace("\u201d", '"')
    s = s.replace("\u2013", "-").replace("\u2014", "-")
    # Normalize unicode (accented e -> e, etc.)
    # But keep the string as NFC so accented chars that we want to preserve stay
    s = unicodedata.normalize("NFC", s)
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_pick_cell(raw):
    """
    Parse a cell like 'Cameron Boozer - Duke' into (name, team|None).
    Handles: 'Name - Team', 'Name -Team', 'Name- Team', 'Name-Team',
             'Name Team' (if Team is a known school), and bare 'Name'.
    """
    raw = normalize_text(raw)
    if not raw:
        return None, None

    # Try splitting on various dash patterns, from most to least specific
    # Use a regex to find a dash that separates name from team
    # Pattern: word-or-punctuation, then dash (possibly with spaces), then team
    # But skip dashes inside compound names like "Wheeler-Thomas", "Kaufman-Renn"
    # Heuristic: if the part after the dash starts with a capital and looks like
    # a school name, it's a separator.

    # Try ' - ' first (clearest separator)
    for sep_re in [r'\s+-\s+', r'\s*-\s+', r'\s+-\s*', r'\s*--\s*']:
        m = re.search(sep_re, raw)
        if m:
            name_part = raw[:m.start()].strip()
            team_part = raw[m.end():].strip()
            if team_part and len(team_part) > 1:
                # Check this isn't a compound name like "Wheeler-Thomas"
                # Team names typically start with uppercase or are abbreviations
                # and don't look like a last name continuation
                team_lower = team_part.lower().rstrip(".")
                # If it's a known team or abbreviation, definitely a team
                if (team_lower in TEAM_OVERRIDES or
                    team_lower in KNOWN_SCHOOLS or
                    len(team_part) <= 4 or  # abbreviation like "BYU", "UCLA", "VCU"
                    team_part[0].isupper()):
                    return name_part, team_part
            # If team_part is empty/short, treat the whole thing as just a name
            if not team_part:
                return name_part, None

    # Try bare dash with no spaces: "Boozer-Duke", "Flemings-Houston"
    # But NOT "Wheeler-Thomas" (compound last name)
    if "-" in raw:
        # Find all dash positions
        parts = raw.split("-")
        if len(parts) == 2:
            name_part = parts[0].strip()
            team_part = parts[1].strip()
            team_lower = team_part.lower().rstrip(".")
            if (team_lower in TEAM_OVERRIDES or
                team_lower in KNOWN_SCHOOLS or
                team_lower in {"south florida", "ohio state", "st johns",
                               "st. johns", "st john's", "iowa st",
                               "michigan st", "north dakota st"}):
                return name_part, team_part
        # For multiple dashes like "Wheeler-Thomas - NDSU", already handled above
        # Fall through to return raw as name

    # Try detecting "Name Team" (no dash) -- e.g., "Darius Acuff Jr. Arkansas"
    # Look for a known school name at the end
    raw_lower = raw.lower()
    for school in sorted(KNOWN_SCHOOLS, key=len, reverse=True):
        if raw_lower.endswith(school):
            # Make sure there's a space before the school name
            prefix = raw[:len(raw) - len(school)].rstrip()
            if prefix and prefix[-1] != "-":
                return prefix.rstrip(". "), raw[len(raw) - len(school):]

    return raw, None


def normalize_name_key(name):
    """Lowercase key for name matching. Strips Jr., suffixes, punctuation."""
    name = name.lower().strip()
    # Remove common suffixes for matching
    name = re.sub(r"\bjr\.?\b", "", name)
    name = re.sub(r"\bsr\.?\b", "", name)
    name = re.sub(r"\bii+\b", "", name)
    name = re.sub(r"\biv\b", "", name)
    # Remove apostrophes, periods, accented chars etc.
    name = name.replace("'", "").replace(".", "").replace("\u2019", "")
    # Normalize accented characters
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = re.sub(r"\s+", " ", name).strip()
    return name


def normalize_team(team):
    """Normalize team name to canonical form."""
    if not team:
        return None
    team = normalize_text(team)
    key = team.lower().strip().rstrip(".")
    if key in TEAM_OVERRIDES:
        return TEAM_OVERRIDES[key]
    # Title-case it
    return team.strip()


def slugify(name):
    """Create a slug from a player name: lowercase, hyphens, no punctuation."""
    s = name.lower().strip()
    s = s.replace("'", "")
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def read_xlsx(path):
    """Read all entrants from the xlsx file."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    entrants = []
    for row in range(2, 100):  # generous upper bound
        name = ws.cell(row=row, column=2).value
        if name is None:
            break
        name = normalize_text(str(name)).strip()

        picks_raw = {}
        for seed in range(1, 17):
            col = seed + 2  # col C=3 is seed 1, ..., col R=18 is seed 16
            val = ws.cell(row=row, column=col).value
            if val:
                picks_raw[seed] = normalize_text(str(val))

        scorer_raw = normalize_text(str(ws.cell(row=row, column=19).value or ""))
        playmaker_raw = normalize_text(str(ws.cell(row=row, column=20).value or ""))

        entrants.append({
            "name": name,
            "picks_raw": picks_raw,
            "scorer_captain_raw": scorer_raw,
            "playmaker_captain_raw": playmaker_raw,
        })

    return entrants


def build_canonical_players(entrants):
    """
    Build canonical player list from all entries.
    Uses the most common / correct spelling of each player.
    Returns dict: normalized_key -> { "name": canonical, "team": team }
    """
    # Collect all (name, team) mentions
    all_mentions = []  # list of (parsed_name, parsed_team, seed)

    for ent in entrants:
        for seed, raw in ent["picks_raw"].items():
            pname, pteam = parse_pick_cell(raw)
            if pname:
                all_mentions.append((pname.strip(), normalize_team(pteam), seed))

    # First pass: apply NAME_OVERRIDES to get canonical names
    # Group by canonical name key
    canonical_groups = {}  # canonical_key -> list of (original, team, seed, canonical_name)

    for pname, pteam, seed in all_mentions:
        key = normalize_name_key(pname)
        # Check overrides
        canonical = NAME_OVERRIDES.get(key, None)
        if not canonical:
            # Try partial last-name-only override
            parts = key.split()
            if parts:
                last = parts[-1]
                canonical = NAME_OVERRIDES.get(last, None)

        if canonical:
            ckey = normalize_name_key(canonical)
        else:
            canonical = pname
            ckey = key

        canonical_groups.setdefault(ckey, []).append((pname, pteam, seed, canonical))

    # For each group, pick the best canonical name and most common team
    canonical_players = {}  # normalized_key -> { name, team, seeds }

    for ckey, mentions in canonical_groups.items():
        # If any mention came from an override, use that override name
        override_names = [m[3] for m in mentions if m[3] != m[0]]
        if override_names:
            best_name = override_names[0]
        else:
            # Pick the most common exact spelling (prefer longer/more detailed ones)
            name_counts = Counter(m[0] for m in mentions)
            candidates = name_counts.most_common()
            # Sort: prefer longer, then more frequent
            candidates.sort(key=lambda x: (-len(x[0]), -x[1]))
            best_name = candidates[0][0]

        # Pick the most common non-None team
        team_counts = Counter(m[1] for m in mentions if m[1])
        best_team = team_counts.most_common(1)[0][0] if team_counts else None

        # If still no team, check PLAYER_TEAMS
        if not best_team and best_name in PLAYER_TEAMS:
            best_team = PLAYER_TEAMS[best_name]

        # Collect all seeds this player was picked at
        seeds = set(m[2] for m in mentions)

        canonical_players[ckey] = {
            "name": best_name,
            "team": best_team,
            "seeds": seeds,
        }

    return canonical_players


def fuzzy_match_player(input_name, canonical_players, seed=None):
    """
    Match an input name to the canonical player list.
    Returns (canonical_key, confidence, match_type) or (None, 0, 'failed').
    """
    key = normalize_name_key(input_name)

    # Check direct override first
    override = NAME_OVERRIDES.get(key)
    if override:
        okey = normalize_name_key(override)
        if okey in canonical_players:
            return okey, 100, "override"

    # Check partial (last name only) override
    parts = key.split()
    if parts:
        last = parts[-1]
        override = NAME_OVERRIDES.get(last)
        if override:
            okey = normalize_name_key(override)
            if okey in canonical_players:
                return okey, 100, "override"

    # Exact key match
    if key in canonical_players:
        return key, 100, "exact"

    # Fuzzy match against all canonical names
    choices = {k: v["name"] for k, v in canonical_players.items()}
    result = process.extractOne(
        input_name,
        choices,
        scorer=fuzz.token_sort_ratio,
    )

    if result:
        matched_name, score, matched_key = result
        if score >= 80:
            return matched_key, score, "fuzzy"
        elif score >= 60:
            return matched_key, score, "low_confidence"

    return None, 0, "failed"


def resolve_captain(captain_raw, entrant_picks, canonical_players):
    """
    Resolve a captain field to (seed, canonical_key).
    Captain field may contain seed number like "(6)" or name + team.
    """
    if not captain_raw:
        return None, None

    raw = captain_raw.strip()

    # Remove trailing seed hint like "(6)" or "(13)"
    seed_hint = None
    m = re.search(r"\((\d+)\)\s*$", raw)
    if m:
        seed_hint = int(m.group(1))
        raw = raw[:m.start()].strip()

    # Parse name from captain field
    pname, pteam = parse_pick_cell(raw)
    if not pname:
        return None, None

    # Try to match to a canonical player
    ckey, conf, mtype = fuzzy_match_player(pname, canonical_players)

    if ckey:
        # Find which seed this player is at in entrant's picks
        for seed, pick_key in entrant_picks.items():
            if pick_key == ckey:
                return seed, ckey

        # If we have a seed hint, use it
        if seed_hint and seed_hint in entrant_picks:
            return seed_hint, entrant_picks[seed_hint]

    # Fallback: if seed hint given, use that seed's player
    if seed_hint and seed_hint in entrant_picks:
        return seed_hint, entrant_picks[seed_hint]

    return None, None


def main():
    print(f"Reading xlsx: {XLSX_PATH}")
    entrants = read_xlsx(XLSX_PATH)
    print(f"Found {len(entrants)} entrants\n")

    # Build canonical player list
    print("=" * 70)
    print("BUILDING CANONICAL PLAYER LIST")
    print("=" * 70)
    canonical_players = build_canonical_players(entrants)

    # Sort by most common seed, then name
    sorted_players = sorted(
        canonical_players.items(),
        key=lambda x: (min(x[1]["seeds"]), x[1]["name"])
    )

    print(f"\n{len(canonical_players)} unique players found:\n")
    print(f"  {'Player':<35} {'Team':<25} Seeds")
    print(f"  {'-'*35} {'-'*25} {'-'*10}")
    for ckey, info in sorted_players:
        seeds_str = ",".join(str(s) for s in sorted(info["seeds"]))
        team = info["team"] or "???"
        print(f"  {info['name']:<35} {team:<25} {seeds_str}")

    # Resolve each entrant
    print(f"\n{'=' * 70}")
    print("RESOLVING ENTRANT PICKS")
    print("=" * 70)

    output_entrants = []
    issues = []

    for ent in entrants:
        print(f"\n  --- {ent['name']} ---")

        resolved_picks = {}  # seed (int) -> canonical_key
        pick_details = {}    # seed (int) -> { player_id, name, team }

        for seed in range(1, 17):
            raw = ent["picks_raw"].get(seed, "")
            if not raw:
                print(f"    Seed {seed:2d}: [MISSING]")
                issues.append(f"{ent['name']}: Missing pick for seed {seed}")
                continue

            pname, pteam = parse_pick_cell(raw)
            if not pname:
                print(f"    Seed {seed:2d}: [EMPTY] '{raw}'")
                issues.append(f"{ent['name']}: Could not parse seed {seed}: '{raw}'")
                continue

            ckey, conf, mtype = fuzzy_match_player(pname, canonical_players)

            if ckey:
                cp = canonical_players[ckey]
                # Always use canonical team
                team = cp["team"] or "Unknown"

                resolved_picks[seed] = ckey
                pick_details[seed] = {
                    "player_id": slugify(cp["name"]),
                    "name": cp["name"],
                    "team": team,
                }

                if mtype == "exact":
                    symbol = "OK"
                elif mtype == "override":
                    symbol = "OVERRIDE"
                elif mtype == "fuzzy":
                    symbol = f"FUZZY({conf})"
                else:
                    symbol = f"LOW({conf})"

                print(f"    Seed {seed:2d}: {symbol:12s} '{pname}' -> {cp['name']} ({team})")
            else:
                print(f"    Seed {seed:2d}: FAILED     '{pname}' -- NO MATCH")
                issues.append(f"{ent['name']}: Seed {seed} '{pname}' - no match found")

        # Resolve captains
        scorer_seed, scorer_key = resolve_captain(
            ent["scorer_captain_raw"], resolved_picks, canonical_players
        )
        playmaker_seed, playmaker_key = resolve_captain(
            ent["playmaker_captain_raw"], resolved_picks, canonical_players
        )

        scorer_captain = None
        playmaker_captain = None

        if scorer_seed and scorer_seed in pick_details:
            scorer_captain = {"seed": scorer_seed, **pick_details[scorer_seed]}
            print(f"    Scorer Captain:    seed {scorer_seed} = {pick_details[scorer_seed]['name']}")
        else:
            print(f"    Scorer Captain:    UNRESOLVED (raw: '{ent['scorer_captain_raw']}')")
            issues.append(f"{ent['name']}: Could not resolve scorer captain '{ent['scorer_captain_raw']}'")

        if playmaker_seed and playmaker_seed in pick_details:
            playmaker_captain = {"seed": playmaker_seed, **pick_details[playmaker_seed]}
            print(f"    Playmaker Captain: seed {playmaker_seed} = {pick_details[playmaker_seed]['name']}")
        else:
            print(f"    Playmaker Captain: UNRESOLVED (raw: '{ent['playmaker_captain_raw']}')")
            issues.append(f"{ent['name']}: Could not resolve playmaker captain '{ent['playmaker_captain_raw']}'")

        # Build output picks dict (string keys)
        picks_out = {}
        for seed in range(1, 17):
            if seed in pick_details:
                picks_out[str(seed)] = pick_details[seed]

        output_entrants.append({
            "name": ent["name"],
            "scorer_captain": scorer_captain,
            "playmaker_captain": playmaker_captain,
            "picks": picks_out,
        })

    # Summary
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    print(f"Entrants: {len(output_entrants)}")
    total_picks = sum(len(e["picks"]) for e in output_entrants)
    print(f"Total picks resolved: {total_picks} / {len(output_entrants) * 16}")
    captains_ok = sum(1 for e in output_entrants if e["scorer_captain"]) + \
                  sum(1 for e in output_entrants if e["playmaker_captain"])
    print(f"Captains resolved: {captains_ok} / {len(output_entrants) * 2}")

    if issues:
        print(f"\nISSUES ({len(issues)}):")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("\nNo issues!")

    # Write output
    output = {
        "year": 2026,
        "entrants": output_entrants,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\nWrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
