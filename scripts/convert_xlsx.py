#!/usr/bin/env python3
"""
Convert the March Madness Fantasy xlsx spreadsheet into picks.json + stats.json.

Usage:
  python convert_xlsx.py <xlsx_file> [--output-dir ../data/2025]
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


def slugify(name):
    """Create a stable ID from a player name."""
    s = name.lower().strip()
    s = re.sub(r"['\.\-]", "", s)
    s = re.sub(r"\s+", "-", s)
    return s


def parse_player_cell(cell_value):
    """
    Parse a player cell like 'Walter Clayton Jr. - Florida' or '*P* Danny Wolf - Michigan'.
    Returns (name, team, is_playmaker_notation).
    """
    if not cell_value or not isinstance(cell_value, str):
        return None, None, False

    val = cell_value.strip()
    is_pm = val.startswith("*P*")
    if is_pm:
        val = val[3:].strip()

    parts = val.rsplit(" - ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip(), is_pm
    return val, "", is_pm


def parse_individual_scores(ws):
    """
    Parse the 'Individual Scores' sheet.
    Returns dict of slug -> { name, team, pts, reb_ast, eliminated, games_by_round }.
    Also returns playmaker_totals: slug -> total (for *P* entries).
    """
    players = {}         # slug -> { name, team, pts, eliminated, rounds }
    playmaker_totals = {} # slug -> total (from *P* entries)
    current_seed = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        a_val = row[0].value  # Column A: seed label
        b_val = row[1].value  # Column B: player name - team
        c_val = row[2].value  # Column C: eliminated?
        d_val = row[3].value  # Column D: total

        if a_val and isinstance(a_val, str) and "seed" in a_val.lower():
            m = re.search(r'(\d+)', a_val)
            if m:
                current_seed = int(m.group(1))
            # Don't continue — the first player of this seed is on the same row

        if not b_val:
            continue

        name, team, is_pm = parse_player_cell(str(b_val))
        if not name:
            continue

        slug = slugify(name)
        total = int(d_val) if d_val and d_val != '' else 0
        eliminated = (str(c_val).strip().upper() == 'X') if c_val else False

        # Parse round-by-round scores (columns E-J)
        rounds = {}
        round_names = ["R64", "R32", "S16", "E8", "F4", "Championship"]
        for i, rname in enumerate(round_names):
            cell = row[4 + i].value if (4 + i) < len(row) else None
            if cell is not None and cell != '':
                try:
                    rounds[rname] = int(float(cell))
                except (ValueError, TypeError):
                    pass

        if is_pm:
            playmaker_totals[slug] = playmaker_totals.get(slug, 0) + total
        elif slug in players:
            # Duplicate entry for same player — accumulate points and rounds
            players[slug]["pts"] += total
            if eliminated:
                players[slug]["eliminated"] = True
            for rname, rpts in rounds.items():
                players[slug]["rounds"][rname] = players[slug]["rounds"].get(rname, 0) + rpts
        else:
            players[slug] = {
                "name": name,
                "team": team,
                "seed": current_seed,
                "pts": total,
                "eliminated": eliminated,
                "rounds": rounds,
            }

    return players, playmaker_totals



# 2025 tournament matchups: team -> { round: opponent }
MATCHUPS_2025 = {
    "Florida": {"R64": "Norfolk State", "R32": "UNC Wilmington", "S16": "Marquette", "E8": "Texas Tech", "F4": "Duke", "Championship": "Houston"},
    "Auburn": {"R64": "Alabama State", "R32": "Yale", "S16": "Michigan", "E8": "Michigan State", "F4": "Houston"},
    "Duke": {"R64": "American", "R32": "Connecticut", "S16": "Arizona", "E8": "Kentucky", "F4": "Florida"},
    "Houston": {"R64": "SIU-Edwardsville", "R32": "Robert Morris", "S16": "Gonzaga", "E8": "Purdue", "F4": "Auburn", "Championship": "Florida"},
    "Alabama": {"R64": "Robert Morris", "R32": "Clemson", "S16": "Illinois", "E8": "Duke"},
    "Tennessee": {"R64": "Lipscomb", "R32": "New Mexico", "S16": "Oregon", "E8": "Florida"},
    "St. John's": {"R64": "Omaha", "R32": "Baylor"},
    "Michigan State": {"R64": "Bryant", "R32": "UC San Diego", "S16": "UCLA", "E8": "Auburn"},
    "Texas Tech": {"R64": "Montana", "R32": "Connecticut", "S16": "Purdue", "E8": "Florida"},
    "Iowa State": {"R64": "Lipscomb", "R32": "Wisconsin"},
    "Wisconsin": {"R64": "Montana", "R32": "Iowa State"},
    "Kentucky": {"R64": "Troy", "R32": "Georgia", "S16": "Marquette", "E8": "Duke"},
    "Arizona": {"R64": "Akron", "R32": "UCF", "S16": "Duke"},
    "Texas A&M": {"R64": "Yale", "R32": "Michigan"},
    "Purdue": {"R64": "High Point", "R32": "Drake", "S16": "Texas Tech", "E8": "Houston"},
    "Maryland": {"R64": "Grand Canyon", "R32": "Gonzaga", "S16": "Tennessee"},
    "Oregon": {"R64": "Liberty", "R32": "Drake", "S16": "Tennessee"},
    "Memphis": {"R64": "Colorado State"},
    "Clemson": {"R64": "McNeese", "R32": "Alabama"},
    "Illinois": {"R64": "Montana", "R32": "Missouri", "S16": "Alabama"},
    "BYU": {"R64": "VCU", "R32": "Creighton", "S16": "Houston"},
    "Marquette": {"R64": "Norfolk State", "R32": "New Mexico", "S16": "Kentucky"},
    "Kansas": {"R64": "Arkansas"},
    "UCLA": {"R64": "Troy", "R32": "Oklahoma", "S16": "Michigan State"},
    "Saint Mary's": {"R64": "Vanderbilt", "R32": "UConn"},
    "Gonzaga": {"R64": "Georgia", "R32": "Maryland", "S16": "Houston"},
    "Louisville": {"R64": "Wofford"},
    "Connecticut": {"R64": "Oklahoma", "R32": "Texas Tech"},
    "Baylor": {"R64": "Mississippi State", "R32": "St. John's"},
    "Creighton": {"R64": "Louisville", "R32": "BYU"},
    "Oklahoma": {"R64": "Connecticut"},
    "Georgia": {"R64": "Gonzaga"},
    "New Mexico": {"R64": "Utah State", "R32": "Marquette"},
    "Arkansas": {"R64": "Kansas", "R32": "Michigan State", "S16": "Michigan"},
    "Vanderbilt": {"R64": "Saint Mary's"},
    "Utah State": {"R64": "New Mexico"},
    "North Carolina": {"R64": "San Diego State"},
    "VCU": {"R64": "BYU"},
    "Drake": {"R64": "Purdue", "R32": "Oregon"},
    "Xavier": {"R64": "Texas"},
    "UC San Diego": {"R64": "Michigan State"},
    "Colorado State": {"R64": "Memphis", "R32": "Tennessee"},
    "Liberty": {"R64": "Oregon"},
    "Yale": {"R64": "Auburn"},
    "Grand Canyon": {"R64": "Maryland"},
    "High Point": {"R64": "Purdue"},
    "Akron": {"R64": "Arizona"},
    "Lipscomb": {"R64": "Tennessee"},
    "Montana": {"R64": "Wisconsin"},
    "Troy": {"R64": "UCLA"},
    "UNC Wilmington": {"R64": "Florida"},
    "Bryant": {"R64": "Michigan State"},
    "Robert Morris": {"R64": "Houston"},
    "Omaha": {"R64": "St. John's"},
    "Wofford": {"R64": "Louisville"},
    "SIU-Edwardsville": {"R64": "Houston"},
    "Norfolk State": {"R64": "Florida"},
    "American": {"R64": "Duke"},
    "Alabama State": {"R64": "Auburn"},
    "Mount St. Mary's": {"R64": "Auburn"},
    "Michigan": {"R64": "UC San Diego", "R32": "Texas A&M", "S16": "Auburn"},
    "Mississippi State": {"R64": "Baylor"},
}


def build_stats_json(players, playmaker_totals):
    """Build stats.json from parsed player data."""
    stats_players = {}

    for slug, p in players.items():
        # Derive reb+ast from playmaker total difference
        reb_ast = 0
        if slug in playmaker_totals:
            reb_ast = playmaker_totals[slug] - p["pts"]

        team = p["team"]
        team_matchups = MATCHUPS_2025.get(team, {})

        games = []
        for rname, pts in p["rounds"].items():
            opponent = team_matchups.get(rname, "")
            games.append({"round": rname, "pts": pts, "reb": 0, "ast": 0, "game_id": "", "opponent": opponent})

        stats_players[slug] = {
            "name": p["name"],
            "team": p["team"],
            "seed": p["seed"],
            "eliminated": p["eliminated"],
            "stats": {
                "pts": p["pts"],
                "reb": reb_ast,  # combined reb+ast (we can't split from sheet data)
                "ast": 0,
            },
            "games": games,
        }

    # Handle players that only exist as *P* entries (no regular entry)
    for slug, pm_total in playmaker_totals.items():
        if slug not in stats_players:
            # This player was only ever used as a playmaker captain
            stats_players[slug] = {
                "name": slug.replace("-", " ").title(),
                "team": "",
                "seed": 0,
                "eliminated": True,
                "stats": {"pts": 0, "reb": pm_total, "ast": 0},
                "games": [],
            }

    return {
        "last_updated": "2025-04-08T00:00:00Z",
        "players": stats_players,
        "active_games": [],
    }


def parse_scoreboard(ws, players_data):
    """
    Parse the 'Sorted Scoreboard' sheet to extract entrant picks.
    """
    entrants = []

    # Column layout (from xlsx):
    # A=name, B=points, C=remaining, D=place
    # E=scorer_captain_seed, F=scorer_captain_player
    # G=playmaker_captain_seed, H=playmaker_captain_player
    # J=seed1_player, K=seed1_pts, L=seed1_elim
    # M=seed2_player, N=seed2_pts, O=seed2_elim
    # ... pattern: every 3 columns for each seed

    # Seed columns start at J (index 9), then M (12), P (15), ...
    # Actually from the data: J=9, M=12, P=15, S=18, V=21, Y=24, AB=27, AE=30, AH=33, AK=36, AN=39, AQ=42, AT=45, AW=48, AZ=51, BC=54
    seed_col_indices = [9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        name = row[0].value  # Column A
        if not name or not isinstance(name, str) or name.strip() == '' or name == '#N/A':
            continue

        total_pts = row[1].value  # Column B
        if total_pts is None:
            continue

        name = name.strip()

        # Captain info
        scorer_seed = int(row[4].value) if row[4].value else None  # Column E
        scorer_player_cell = row[5].value  # Column F
        playmaker_seed = int(row[6].value) if row[6].value else None  # Column G
        playmaker_player_cell = row[7].value  # Column H

        scorer_name, scorer_team, _ = parse_player_cell(scorer_player_cell)
        playmaker_name, playmaker_team, _ = parse_player_cell(playmaker_player_cell)

        # Build picks for each seed
        picks = {}
        for seed_num in range(1, 17):
            col_idx = seed_col_indices[seed_num - 1]
            if col_idx >= len(row):
                continue

            player_cell = row[col_idx].value
            if not player_cell:
                continue

            pname, pteam, is_pm = parse_player_cell(str(player_cell))
            if not pname:
                continue

            slug = slugify(pname)
            picks[str(seed_num)] = {
                "player_id": slug,
                "name": pname,
                "team": pteam,
            }

        # Build captain entries
        scorer_captain = None
        if scorer_seed and scorer_name:
            scorer_slug = slugify(scorer_name)
            scorer_captain = {
                "seed": scorer_seed,
                "player_id": scorer_slug,
                "name": scorer_name,
                "team": scorer_team,
            }

        playmaker_captain = None
        if playmaker_seed and playmaker_name:
            playmaker_slug = slugify(playmaker_name)
            playmaker_captain = {
                "seed": playmaker_seed,
                "player_id": playmaker_slug,
                "name": playmaker_name,
                "team": playmaker_team,
            }

        entrants.append({
            "name": name,
            "scorer_captain": scorer_captain,
            "playmaker_captain": playmaker_captain,
            "picks": picks,
            "_expected_total": float(total_pts),  # for verification
        })

    return entrants


def verify_scores(entrants, stats_data):
    """Verify computed scores match spreadsheet totals."""
    print("\nVerifying scores...")
    mismatches = 0

    for entrant in entrants:
        total = 0
        for seed in range(1, 17):
            pick = entrant["picks"].get(str(seed))
            if not pick:
                continue

            pid = pick["player_id"]
            player = stats_data["players"].get(pid)
            if not player:
                print(f"  Warning: {entrant['name']} seed {seed} - player '{pid}' not in stats")
                continue

            pts = player["stats"]["pts"]
            reb = player["stats"]["reb"]
            ast = player["stats"]["ast"]

            if entrant.get("scorer_captain") and entrant["scorer_captain"]["player_id"] == pid:
                total += pts * 1.5
            elif entrant.get("playmaker_captain") and entrant["playmaker_captain"]["player_id"] == pid:
                total += pts + reb + ast
            else:
                total += pts

        expected = entrant.get("_expected_total", 0)
        diff = abs(total - expected)
        status = "✓" if diff < 0.1 else f"✗ (diff={diff:.1f})"
        if diff >= 0.1:
            mismatches += 1
        print(f"  {entrant['name']:25s}: computed={total:7.1f}  expected={expected:7.1f}  {status}")

    if mismatches:
        print(f"\n⚠ {mismatches} mismatches found!")
    else:
        print(f"\n✓ All {len(entrants)} entrants match!")

    return mismatches


def main():
    parser = argparse.ArgumentParser(description="Convert xlsx to picks.json + stats.json")
    parser.add_argument("xlsx_file", help="Path to the xlsx file")
    parser.add_argument("--output-dir", default=None, help="Output directory")
    parser.add_argument("--year", type=int, default=2025)
    args = parser.parse_args()

    if not args.output_dir:
        args.output_dir = str(Path(__file__).parent.parent / "data")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Loading {args.xlsx_file}...")
    wb = openpyxl.load_workbook(args.xlsx_file, data_only=True)

    # Parse individual scores
    print("\nParsing Individual Scores...")
    players, playmaker_totals = parse_individual_scores(wb["Individual Scores"])
    print(f"  {len(players)} regular players, {len(playmaker_totals)} playmaker entries")

    # Build stats.json
    stats_data = build_stats_json(players, playmaker_totals)
    print(f"  {len(stats_data['players'])} total players in stats")

    # Parse scoreboard
    print("\nParsing Sorted Scoreboard...")
    entrants = parse_scoreboard(wb["Sorted Scoreboard"], players)
    print(f"  {len(entrants)} entrants")

    # Verify
    mismatches = verify_scores(entrants, stats_data)

    # Remove verification field before writing
    for e in entrants:
        e.pop("_expected_total", None)

    # Write files
    picks_data = {"year": args.year, "entrants": entrants}
    stats_path = output_dir / "stats.json"
    picks_path = output_dir / "picks.json"

    with open(picks_path, "w") as f:
        json.dump(picks_data, f, indent=2)
    print(f"\nWrote {picks_path}")

    with open(stats_path, "w") as f:
        json.dump(stats_data, f, indent=2)
    print(f"Wrote {stats_path}")

    # Also copy to 2025 archive
    archive_dir = output_dir / "2025"
    archive_dir.mkdir(parents=True, exist_ok=True)
    with open(archive_dir / "picks.json", "w") as f:
        json.dump(picks_data, f, indent=2)
    with open(archive_dir / "stats.json", "w") as f:
        json.dump(stats_data, f, indent=2)
    print(f"Wrote 2025 archive to {archive_dir}")

    if mismatches:
        print(f"\n⚠ {mismatches} score mismatches - review above!")
        sys.exit(1)


if __name__ == "__main__":
    main()
